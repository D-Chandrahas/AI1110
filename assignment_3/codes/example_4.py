from openpyxl import load_workbook


def find_prob(item,dist):
    return dist[item]/sum(dist.values())

if __name__ == "__main__":

    ws = load_workbook("prob_dist.xlsx").active

    for items in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        pass

    for freq in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        pass
    
    dist = dict((x,y) for x,y in zip(items,freq))
        
    print(find_prob(6,dist))
    
    